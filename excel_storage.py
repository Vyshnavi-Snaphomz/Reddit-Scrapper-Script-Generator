import os
import json
from datetime import datetime
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook


EXCEL_PATH = os.getenv("SCRAPED_EXCEL_PATH", "scraped_data.xlsx")
SHEET_NAME = "scraped_data"
HEADERS = [
    "subreddit",
    "post",
    "posted on",
    "author",
    "likes",
    "Comments",
    "post_age_days",
    "post_flair",
    "post_id",
    "post_title",
    "post_selftext",
    "post_url",
    "post_upvote_ratio",
    "post_num_comments",
    "post_rank",
    "comment_id_array",
    "comment_author_array",
    "comment_body_array",
    "comment_ups_array",
    "comment_url_array",
    "comment_created_utc_array",
    "comments_array",
    "imgbb_link",
    "scraped_at_utc",
]


def _now_iso() -> str:
    return datetime.utcnow().isoformat()


def _dt_to_iso(val):
    return val.isoformat() if val else None


def _open_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active if wb.sheetnames else wb.create_sheet(SHEET_NAME)
        ws.title = SHEET_NAME

    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.delete_rows(1, 1)

    if ws.max_row == 0:
        ws.append(HEADERS)
    else:
        existing = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
        if existing != HEADERS:
            for idx, header in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=idx, value=header)
    return wb


def _save_workbook(wb) -> None:
    wb.save(EXCEL_PATH)


def _sheet_rows_as_dicts(ws, headers: List[str]) -> List[Dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows


def _append_rows(rows: Iterable[List]):
    wb = _open_workbook()
    ws = wb[SHEET_NAME]
    for row in rows:
        ws.append(row)
    _save_workbook(wb)


def append_subreddit_block(
    subreddit: str,
    query: Optional[str] = None,
    title: Optional[str] = None,
    description: Optional[str] = None,
) -> None:
    # Keep this for compatibility; write a full-width row.
    _append_rows(
        [
            [
                subreddit,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                _now_iso(),
            ]
        ]
    )


def append_post_comment_block(post: Dict, comments: Iterable[Dict], show_subreddit: bool) -> None:
    post_rank = post.get("post_rank")
    post_label = f"post{post_rank}" if post_rank else "post"
    comment_rows = list(comments) if comments else []
    comments_array = [
        {
            "comment_id": c.get("id"),
            "author": c.get("author"),
            "body": c.get("body"),
            "ups": c.get("ups"),
            "url": c.get("url"),
            "created_utc": _dt_to_iso(c.get("created_utc")),
        }
        for c in comment_rows
    ]
    comment_id_array = [c.get("comment_id") for c in comments_array]
    comment_author_array = [c.get("author") for c in comments_array]
    comment_body_array = [c.get("body") for c in comments_array]
    comment_ups_array = [c.get("ups") for c in comments_array]
    comment_url_array = [c.get("url") for c in comments_array]
    comment_created_utc_array = [c.get("created_utc") for c in comments_array]

    row = [
        post.get("subreddit") if show_subreddit else None,
        post_label,
        _dt_to_iso(post.get("created_utc")),
        post.get("author"),
        post.get("ups"),
        len(comment_rows),
        post.get("post_age_days"),
        post.get("link_flair_text"),
        post.get("id"),
        post.get("title"),
        post.get("selftext"),
        post.get("url"),
        post.get("upvote_ratio"),
        post.get("num_comments"),
        post_rank,
        json.dumps(comment_id_array, ensure_ascii=False),
        json.dumps(comment_author_array, ensure_ascii=False),
        json.dumps(comment_body_array, ensure_ascii=False),
        json.dumps(comment_ups_array, ensure_ascii=False),
        json.dumps(comment_url_array, ensure_ascii=False),
        json.dumps(comment_created_utc_array, ensure_ascii=False),
        json.dumps(comments_array, ensure_ascii=False),
        post.get("imgbb_link"),
        _now_iso(),
    ]
    _append_rows([row])


def append_post_row(post: Dict, post_rank: int) -> None:
    # Backward compatibility no-op; writing happens with comments for sample layout.
    return None


def append_comment_rows(subreddit: str, post: Dict, comments: Iterable[Dict]) -> None:
    append_post_comment_block(post, comments, show_subreddit=False)


def append_subreddits(records: Iterable[Dict]) -> None:
    for rec in records:
        append_subreddit_block(
            subreddit=rec.get("name"),
            query=rec.get("query"),
            title=rec.get("title"),
            description=rec.get("description"),
        )


def append_posts(records: Iterable[Dict]) -> None:
    return None


def append_comments(subreddit: str, post_id: str, comments: Iterable[Dict]) -> None:
    return None


def _all_rows() -> List[Dict]:
    wb = _open_workbook()
    ws = wb[SHEET_NAME]
    return _sheet_rows_as_dicts(ws, HEADERS)


def get_subreddits(query: Optional[str] = None) -> List[str]:
    rows = _all_rows()
    out = []
    seen = set()
    for row in rows:
        subreddit = row.get("subreddit")
        if not subreddit:
            continue
        if subreddit in seen:
            continue
        seen.add(subreddit)
        out.append(subreddit)
    return out


def counts() -> Dict[str, int]:
    rows = _all_rows()
    total_comments = 0
    for r in rows:
        try:
            arr = json.loads(r.get("comments_array") or "[]")
            total_comments += len(arr) if isinstance(arr, list) else 0
        except Exception:
            continue
    return {
        "rows": len(rows),
        "subreddits": len({r.get("subreddit") for r in rows if r.get("subreddit")}),
        "posts": len({r.get("post_id") for r in rows if r.get("post_id")}),
        "comments": total_comments,
    }


def get_recent_rows(limit: int = 50) -> List[Dict]:
    rows = _all_rows()
    if limit <= 0:
        return rows
    return rows[-limit:]


def get_all_rows() -> List[Dict]:
    return _all_rows()
